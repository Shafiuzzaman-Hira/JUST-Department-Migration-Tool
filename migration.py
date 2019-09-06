import logging
import os
from datetime import datetime

import pyodbc
from openpyxl import Workbook

# Connection String
connection_string: object = pyodbc.connect(
    'Driver={SQL Server};'
    'Server=.\sqlexpress;'
    'Database=Admission2019;'
    'Trusted_Connection=yes;')
cursor: object = connection_string.cursor()

# Excel
wb = Workbook(write_only=True)
migration_ws = wb.create_sheet("Migration")
department_ws = wb.create_sheet("Department")


def backup_db() -> object:
    print("Database backup started...")
    logging.info("Database backup started...")
    connection_string.autocommit = True
    backup_file = "'db.bak'"
    sql = "BACKUP DATABASE [Admission2019] TO DISK = N" + backup_file
    cursor.execute(sql)
    connection_string.autocommit = False
    print("Find the backup file in " + backup_file)
    logging.info("Find the backup file in " + backup_file)
    print("Database backup finished...")
    logging.info("Database backup finished...")


# Getting Applicants who did not cancelled admission and did not initiate Auto Migration OFF of Unit
def get_applicants(unit):
    print("Getting Applicants who did not cancelled admission and did not initiate Auto Migration OFF of Unit " + unit)
    logging.info(
        "Getting Applicants who did not cancelled admission and did not initiate Auto Migration OFF of Unit " + unit)
    cursor.execute(
        f"SELECT * FROM [PassedApplicants] WHERE IsAdmissionCancelled != 1 and [IsAutoMigrationOff] != 1 and UnitName ='{unit}'")
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


# get starting position of migration
def get_first_position(unit):
    logging.info("Getting 1st Applicant...")
    cursor.execute(
        'SELECT MIN(Position) FROM PassedApplicants WHERE [IsAdmissionCancelled] != 1 and [IsAutoMigrationOff] != 1 and UnitName =\'%s\'' % unit)
    for row in cursor.fetchall():
        return row[0]


def get_subject_choices_by_id(application_id: object) -> object:
    cursor.execute(
        "SELECT SubjectId, [Order], ApplicationId FROM Admission2019.SubjectChoices WHERE ApplicationId ='%s'" % application_id)
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def get_subject_id_by_order(application_table_id, order):
    cursor.execute("SELECT SubjectId FROM Admission2019.SubjectChoices WHERE ApplicationId = ? AND [Order] = ?",
                   application_table_id, order)
    for row in cursor.fetchall():
        return row[0]


def get_department_status_by_id(department_id):
    cursor.execute(
        "SELECT SeatStatus, TotalSeats, AllottedSeats, DepartmentName FROM Departments WHERE Id =?", department_id)
    for row in cursor.fetchall():
        return row[0], row[1], row[2], row[3]


def get_departments():
    cursor.execute("SELECT * FROM Departments WHERE SeatStatus = 1")
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def allocate_subject(application_table_id, applicant_position):
    choices = get_subject_choices_by_id(application_table_id)
    number_of_choices = len(choices)
    logging.info("No. of Choices: " + str(number_of_choices))
    x = datetime.now()
    # Make IsAdmissionCancelled True for the applicant who did not fill up the choice form
    if number_of_choices == 0:
        cursor.execute("UPDATE [PassedApplicants] SET [IsAdmissionCancelled] = ?, [UpdatedDate] = ? WHERE [Id] = ?",
                       1, x.strftime("%d%b%I%M%p"), application_table_id)
        logging.info("Admission is being cancelled as ")
        return "did not fill up the choice form"
    order = 1
    while order <= number_of_choices:
        subject_id = get_subject_id_by_order(application_table_id, order)
        seat_status, total_seats, allotted_seats, department_name = get_department_status_by_id(subject_id)
        logging.info(
            str(order) + ": " + str(department_name) + " Total Seats: " + str(total_seats) + " Allotted Seats: " + str(
                allotted_seats))
        if seat_status is True and allotted_seats < total_seats and total_seats is not 0:
            cursor.execute("UPDATE [PassedApplicants] SET [AllottedDepartment] = ?, [UpdatedDate] = ? WHERE [Id] = ?",
                           department_name, x.strftime("%d%b%I%M%p"), application_table_id)
            allotted_seats = allotted_seats + 1
            cursor.execute("UPDATE Departments SET [AllottedSeats] = ? WHERE [Id] = ?",
                           allotted_seats, subject_id)
            if allotted_seats == 1:
                cursor.execute("UPDATE Departments SET [AllottedSeats] = ?, [StartingPosition] = ? WHERE [Id] = ?",
                               allotted_seats, applicant_position, subject_id)
            if allotted_seats == total_seats:
                cursor.execute(
                    "UPDATE Departments SET [EndingPosition] = ?, [SeatStatus] = ? WHERE [Id] = ?",
                    applicant_position, 0, subject_id)
            return department_name
        else:
            order = order + 1
    return "No Department"


def get_applicant_id_by_position(applicant_position):
    cursor.execute("SELECT Id FROM PassedApplicants WHERE Position = '%d'" % applicant_position)
    for row in cursor.fetchall():
        return row[0]


def write_migration_data_to_excel(applicants_data):
    logging.info("Writing migration data to excel......")
    print("Writing migration data to excel......")
    # write header
    migration_ws.append(["Position", "Name", "ApplicationId", "Roll", "Department", "Unit", "Phone", "Quota"
                                                                                                     "IsAutoMigrationOff"])
    # write data
    for applicant in applicants_data:
        position = applicant[1]
        name = applicant[7]
        phone = applicant[6]
        applicant_id = applicant[2]
        roll = applicant[3]
        department = applicant[11]
        unit = applicant[4]
        quota = applicant[10]
        auto_migration_off = applicant[13]
        migration_ws.append([position, name, applicant_id, roll, department, unit, phone, quota, auto_migration_off])


def write_department_data_to_excel(department_data):
    print("Writing department to excel......")
    logging.info("Writing department to excel......")
    # write header
    department_ws.append(
        ["Department", "Unit", "Total Seats", "Allotted Seats", "Seat Status", "Starting Position", "Ending Position"])
    # write data
    for department in department_data:
        department_ws.append(
            [department[2], department[3], department[4], department[5], department[6], department[7], department[8]])


if __name__ == '__main__':
    drive = input("Directory Name (Ex: C/D): ")
    os.chdir(drive + ':')  # change directory
    path: object = "Migration" + str(datetime.now().strftime("_%d_%B_%I_%M_%p"))
    if not os.path.exists(path):
        os.makedirs(path)
    os.chdir(path)

    logging.basicConfig(filename="info.log", format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S',
                        level=logging.INFO)

    unit_name = input("Unit(Press A/B/C/D/E/F): ")
    unit_name = unit_name.upper()
    if unit_name == 'A' or unit_name == 'B' or unit_name == 'C' or unit_name == 'D' or unit_name == 'E' or unit_name == 'F':
        while True:
            try:
                end_position = int(input('End Position: '))
                break
            except:
                print("Invalid Input!")

        backup_db()
        applicants = get_applicants(unit_name)
        position = get_first_position(unit_name)  # get starting position of migration
        logging.info('This is an info message')
        logging.info("Starting Position: " + str(position))
        no_of_applicants = len(applicants)
        logging.info("Total: " + str(no_of_applicants))
        while position <= end_position:
            logging.info(
                "---------------------------------" + str(no_of_applicants) + " remains ------------------------------")
            logging.info("Migration Started for Position " + str(position))
            applicant_id = get_applicant_id_by_position(position)
            logging.info("Id: " + str(applicant_id))
            if applicant_id is not None:
                department = allocate_subject(applicant_id, position)
                logging.info("Position " + str(position) + " " + department)
                cursor.commit()  # Commit db changes
            position = position + 1
            no_of_applicants = no_of_applicants - 1
            if no_of_applicants == 0:
                break
        logging.info("Exporting migration result into excel....")
        print("Exporting migration result into excel....")
        get_applicants_query = " SELECT * FROM PassedApplicants WHERE IsAdmissionCancelled = 0 "
        migration_data = cursor.execute(get_applicants_query)
        write_migration_data_to_excel(migration_data)
        logging.info("Exporting department status into excel....")
        print("Exporting department status into excel....")
        get_departments_query = " SELECT * FROM Departments"
        department_data = cursor.execute(get_departments_query)
        write_department_data_to_excel(department_data)

        migration_result = "Migration.xlsx"
        wb.save(migration_result)
        logging.info("Find excel file into " + migration_result)
        print("Find excel file into " + migration_result)
    else:
        logging.warning("Invalid Unit Name: " + unit_name)
        print("Invalid Unit Name")
